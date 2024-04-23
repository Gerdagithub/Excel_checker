# -*- coding: utf-8 -*-
"""
Created on Tue Apr 23 20:02:41 2024

@author: gerda
"""

import openpyxl
import os
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from constants import *
from custom_logging import *
from files import *

def get_workbook(file_path):    
    try:
        wb_values = openpyxl.load_workbook(file_path, data_only=True)
        wb_formulas = openpyxl.load_workbook(file_path, data_only=False)

        sheets_num = len(wb_values.sheetnames)
        file_name = os.path.basename(file_path)
        return {WORKBOOK_WITH_CALCULATED_VALUES: wb_values, 
                WORKBOOK_WITH_FORMULAS: wb_formulas}
        
    except FileNotFoundError:
        logging.error(f"File not found. Check the path: {file_path}")
        
    except Exception as e:
        logging.exception("An unexpected error occurred")
        
    return None


'''
Returns response about a cell. 
If 'None' is returned, it's not a formula or it's identical
'''
def check_cell(ans_val, ans_fmla, sol_val, sol_fmla, cell_index, ws_title):
    # Answer's cell is empty - it's okay
    if ((ans_val is None) and (ans_fmla is None)):
        return None
    
    # Answer's cell is not empty but student's cell is empty
    if ((ans_val is not None) and (ans_fmla is not None) and
        ((sol_val is None) or (sol_fmla is None))):
        
        return {'type': LOG_TYPE_WARNING, 
                'value': 
                f"Worksheet: {ws_title}\n" + 
                f"Cell is empty instead of being filled {cell_index}\n\n" +
                f"{'Actual value:'.ljust(20)} None\n" +
                f"{'Expected value:'.ljust(20)} {ans_val}\n\n" +
                f"{'Actual formula:'.ljust(20)} None\n" +
                f"{'Expected formula:'.ljust(20)} {ans_fmla}\n"}
    
    # Calculated values are not equal
    if (ans_val != sol_val): 
        return {'type': LOG_TYPE_WARNING, 
                'value': 
                f"Worksheet: {ws_title}\n" +
                f"Invalid value of a cell {cell_index}\n\n" +
                f"{'Actual value:'.ljust(20)} {sol_val}\n" +
                f"{'Expected value:'.ljust(20)} {ans_val}\n\n" +
                f"{'Actual formula:'.ljust(20)} {sol_fmla}\n" +
                f"{'Expected formula:'.ljust(20)} {ans_fmla}\n"}
    
    # If this cell is not formula, return
    if (isinstance(ans_fmla, str) and (not ans_fmla.startswith("="))):
        return None
    
    # If this cell should be formula, but student did something different  
    if ((isinstance(ans_fmla, str) and (ans_fmla.startswith("="))) and 
        ((not isinstance(sol_fmla, str)) or (not sol_fmla.startswith("=")))):
 
        return {'type': LOG_TYPE_WARNING, 
                'value': 
                f"Worksheet: {ws_title}\n" +
                f"Not a formula in cell {cell_index}\n\n" +
                f"{'Actual value:'.ljust(20)} {sol_val}\n" +
                f"{'Expected value:'.ljust(20)} {ans_val}\n\n" +
                f"{'Actual formula:'.ljust(20)} {sol_fmla}\n" +
                f"{'Expected formula:'.ljust(20)} {ans_fmla}\n"}
       
    # If actual and student's formulas don't match 
    # (just printing it because teacher needs to check)
    if (ans_fmla != sol_fmla):
        return {'type': LOG_TYPE_INFO, 
                'value': 
                f"Worksheet: {ws_title}\n" +
                f"Not a formula in cell {cell_index}\n\n" +
                f"{'Actual value:'.ljust(20)} {sol_val}\n" +
                f"{'Expected value:'.ljust(20)} {ans_val}\n\n" +
                f"{'Actual formula:'.ljust(20)} {sol_fmla}\n" +
                f"{'Expected formula:'.ljust(20)} {ans_fmla}\n"}
    
    return None


def check_solutions_of_sheet(answers_ws_vals, answers_ws_fmls, solutions_ws_vals, 
                             solutions_ws_fmls, max_row=400, max_col=20):
    
    if (not (isinstance(answers_ws_vals, Worksheet)) and 
        (isinstance(answers_ws_fmls, Worksheet)) and 
        (isinstance(solutions_ws_vals, Worksheet)) and 
        (isinstance(solutions_ws_fmls, Worksheet))):
        return False
    
    responses = []
    global responses_number
    responses_number = 0
        
    # Compare cells in the defined range
    for row in range(1, max_row + 1):
        for col in range(1, max_col + 1):
            col_name = chr(64 + col)
            cell_index = f"{chr(64 + col)}{row}"  # Generates A1, B1, ..., J10, etc.
            
            sol_val = solutions_ws_vals[cell_index].value
            sol_fmla = solutions_ws_fmls[cell_index].value
            ans_val = answers_ws_vals[cell_index].value
            ans_fmla = answers_ws_fmls[cell_index].value
            ws_title = solutions_ws_vals.title
            
            response = check_cell(ans_val, ans_fmla, sol_val, 
                                  sol_fmla, cell_index, ws_title)
            
            if (response == None): # Student and teacher cells match
                continue
            
            col_count = sum(1 for response in responses if response['column'] == col_name)
            if (col_count >= MAX_SAME_COLUMN_RESPONSES):
                continue
            
            responses.append({
                'column': col_name, 
                'response': response})
            
            
            
            # print_log_message(response)
    if (len(responses) == 0):
        return True
    
    logging.info("-------------------------------- " + answers_ws_vals.title + 
                 " --------------------------------")        

    log_responses(responses)
    return True # Function executed succesfully


def check_solutions_of_wb(answers_wb, solutions_wbs): 
    for sol_wb_path in solutions_wbs:
        logging.info(SEPARATOR)
        logging.info("Checking %s\n", sol_wb_path)
        sol_wb = get_workbook(os.path.join(SOLUTIONS_PATH, sol_wb_path))
        for ans_ws_val, ans_ws_fmls, sol_ws_val, sol_ws_fmls, in zip(
            answers_wb.get(WORKBOOK_WITH_CALCULATED_VALUES).worksheets, 
            answers_wb.get(WORKBOOK_WITH_FORMULAS).worksheets,
            sol_wb.get(WORKBOOK_WITH_CALCULATED_VALUES).worksheets,
            sol_wb.get(WORKBOOK_WITH_FORMULAS).worksheets):
            
            
            status = check_solutions_of_sheet(ans_ws_val, ans_ws_fmls, 
                                              sol_ws_val, sol_ws_fmls)
            if (status == False):
                print(f"Something went wrong in {sol_wb_path}")
                continue
            
            
            
            
            
            